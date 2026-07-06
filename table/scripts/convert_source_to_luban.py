# table/scripts/convert_source_to_luban.py
# One-shot: existing source/*.xlsx (row1=types, row2=names, row3+=data)
#   -> Luban Excel-embedded Datas/#Name.xlsx (##var/##type/##group/## + data)
#   -> Datas/__tables__.xlsx (read_schema_from_file=TRUE)
import os
from openpyxl import load_workbook, Workbook

SRC = os.path.join(os.path.dirname(__file__), "..", "source")
OUT = os.path.join(os.path.dirname(__file__), "..", "Datas")

# Per-table config:
#   value_type : Luban bean class name (PascalCase)
#   index      : key field (snake_case column in source)
#   table_group: "" => all groups; "c" => whole table client-only
#   field_groups: {column_name: "c"} per-field overrides (blank = all groups)
TABLES = {
    "Character": {"value_type": "Character", "index": "code", "table_group": "",
                  "field_groups": {"description": "c"}},
    "Skin":      {"value_type": "Skin", "index": "code", "table_group": "",
                  "field_groups": {"description": "c"}},
    "SkinAsset": {"value_type": "SkinAsset", "index": "code", "table_group": "c",
                  "field_groups": {}},
    "Item":      {"value_type": "Item", "index": "code", "table_group": "",
                  "field_groups": {"description": "c"}},
    "StatusEffect": {"value_type": "StatusEffect", "index": "id", "table_group": "",
                     "field_groups": {"description": "c"}},
    # Ability는 다형 effects 리스트(타입별 AbilityEffect)라 평면 흐름 밖이다.
    #   - effect bean 계층 = write_beans_index()가 __beans__에 정의
    #   - TbAbility 등록 = write_tables_index()가 명시 append
    #   - 데이터(#Ability.xlsx) = Luban embedded 형식으로 수기 author(평면 source 없음)
}

# C# reserved keywords cannot be Luban field names. Normalize per table.
FIELD_RENAME = {
}

def read_source(name):
    wb = load_workbook(os.path.join(SRC, f"{name}.xlsx"), data_only=True)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    types = rows[0]
    names = rows[1]
    # trim trailing all-None columns
    ncol = max((i + 1 for i, n in enumerate(names) if n not in (None, "")), default=0)
    types = [("" if t is None else str(t)) for t in types[:ncol]]
    names = [("" if n is None else str(n)) for n in names[:ncol]]
    data = [[("" if v is None else v) for v in r[:ncol]] for r in rows[2:]
            if any(v not in (None, "") for v in r[:ncol])]
    return names, types, data

def write_table(name, cfg):
    names, types, data = read_source(name)
    rename = FIELD_RENAME.get(name, {})
    names = [rename.get(n, n) for n in names]
    groups = [cfg["field_groups"].get(n, "") for n in names]
    wb = Workbook(); ws = wb.active
    ws.append(["##var"]   + names)
    ws.append(["##type"]  + types)
    ws.append(["##group"] + groups)
    ws.append(["##"]      + names)        # human comment row
    for row in data:
        ws.append([""] + list(row))       # data rows: col A empty
    wb.save(os.path.join(OUT, f"#{name}.xlsx"))

def write_tables_index():
    wb = Workbook(); ws = wb.active
    # Luban 4.9.0's built-in __TableRecord__ bean REQUIRES the `output` column
    # (and accepts `tags`). Canonical header:
    #   full_name | value_type | read_schema_from_file | input | index | mode | group | comment | tags | output
    ws.append(["##var", "full_name", "value_type", "read_schema_from_file",
               "input", "index", "mode", "group", "comment", "tags", "output"])
    for name, cfg in TABLES.items():
        ws.append(["", f"Tb{cfg['value_type']}", cfg["value_type"], "TRUE",
                   f"#{name}.xlsx", cfg["index"], "map", cfg["table_group"], name, "", ""])
    # Ability: 다형 effects 리스트 → 평면 흐름 밖. #Ability.xlsx는 수기 author(Luban embedded).
    ws.append(["", "TbAbility", "Ability", "TRUE", "#Ability.xlsx", "id", "map", "", "Ability", "", ""])
    wb.save(os.path.join(OUT, "__tables__.xlsx"))

def write_beans_index():
    """
    __beans__.xlsx — 다형 AbilityEffect 계층 정의.
    어빌리티의 effects 리스트가 담는 타입들(base + 구체 effect). 자식이 parent를 가리키면 base는 자동 abstract.
    필드는 `*fields` 멀티컬럼 그룹(서브헤더 name/alias/type/...)에 인라인 — 헤더 셀은 *반드시 병합*(J1:P1).
    필드가 여러 개인 bean(DamageEffect)은 첫 행에 full_name + 첫 필드, 이어지는 행은 full_name 비우고
    필드만(같은 bean의 *fields 리스트에 이어붙음 = Luban 다행 연속 표기).
    """
    wb = Workbook(); ws = wb.active
    ws.append(["##var", "full_name", "parent", "valueType", "alias", "sep",
               "comment", "tags", "group", "*fields", "", "", "", "", "", ""])
    ws.append(["##var", "", "", "", "", "", "", "", "",
               "name", "alias", "type", "group", "comment", "tags", "variants"])
    ws.append(["", "AbilityEffect", "", "", "", "", "polymorphic base", "", "",
               "", "", "", "", "", "", ""])
    ws.append(["", "StatusEffectApplyEffect", "AbilityEffect", "", "", "", "", "", "",
               "status_effect_id", "", "int", "", "", "", ""])
    ws.append(["", "MotionEffect", "AbilityEffect", "", "", "", "", "", "",
               "speed", "", "float", "", "", "", ""])
    # DamageEffect: 다중 필드(amount/range/angle). 첫 필드만 본 행, 나머지는 연속 행.
    ws.append(["", "DamageEffect", "AbilityEffect", "", "", "", "데미지 + 부채꼴 타게팅 형상", "", "",
               "amount", "", "int", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "range", "", "float", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "angle", "", "float", "", "", "", ""])
    # KnockbackEffect: 다중 필드(strength/range/angle/duration_ticks/decay_per_tick). DamageEffect와 동형.
    ws.append(["", "KnockbackEffect", "AbilityEffect", "", "", "", "넉백 — 공격자 반대로 밀기(지수 감쇠)", "", "",
               "strength", "", "float", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "range", "", "float", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "angle", "", "float", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "duration_ticks", "", "int", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "",
               "decay_per_tick", "", "float", "", "", "", ""])
    ws.merge_cells("J1:P1")
    wb.save(os.path.join(OUT, "__beans__.xlsx"))

def write_enums_index():
    """Write an empty __enums__.xlsx with the required Luban 4.9.0 column headers."""
    wb = Workbook(); ws = wb.active
    # Luban 4.9.0 ExcelSchemaLoader.LoadEnumListFromFile required columns:
    ws.append(["##var", "full_name", "comment", "flags", "group",
               "tags", "unique"])
    wb.save(os.path.join(OUT, "__enums__.xlsx"))

def main():
    os.makedirs(OUT, exist_ok=True)
    for name, cfg in TABLES.items():
        write_table(name, cfg)
        print(f"[OK] Datas/#{name}.xlsx")
    write_tables_index();  print("[OK] Datas/__tables__.xlsx")
    write_beans_index();   print("[OK] Datas/__beans__.xlsx")
    write_enums_index();   print("[OK] Datas/__enums__.xlsx")
    print("[DONE]")

if __name__ == "__main__":
    main()
